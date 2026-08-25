import io
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..auth import require_hr
from ..database import get_db
from ..models import Department, Employee, Position
from ..schemas import EmployeeIn, EmployeeOut, EmployeeUpdate

router = APIRouter(prefix="/employees", tags=["员工管理"], dependencies=[Depends(require_hr)])

EXPORT_HEADERS = ["工号", "姓名", "性别", "手机号", "邮箱", "部门", "职位", "入职日期", "状态"]


def emp_out(emp: Employee) -> EmployeeOut:
    return EmployeeOut(
        id=emp.id,
        name=emp.name,
        employee_no=emp.employee_no,
        gender=emp.gender,
        phone=emp.phone,
        email=emp.email,
        department_id=emp.department_id,
        department_name=emp.department.name if emp.department else "",
        position_id=emp.position_id,
        position_name=emp.position.name if emp.position else "",
        hire_date=emp.hire_date,
        status=emp.status,
    )


def build_query(db: Session, keyword: str = "", department_id: int | None = None,
                status: str = ""):
    q = db.query(Employee)
    if keyword:
        q = q.filter(
            (Employee.name.like(f"%{keyword}%")) | (Employee.employee_no.like(f"%{keyword}%"))
        )
    if department_id:
        q = q.filter(Employee.department_id == department_id)
    if status:
        q = q.filter(Employee.status == status)
    return q


# 注意：/all 与 /export/excel 必须定义在 /{emp_id} 之前，避免被路径参数吞掉
@router.get("", response_model=dict)
def list_employees(keyword: str = "", department_id: int | None = None,
                   status: str = "", page: int = 1, size: int = 10,
                   db: Session = Depends(get_db)):
    q = build_query(db, keyword, department_id, status)
    total = q.count()
    items = q.order_by(Employee.id).offset((page - 1) * size).limit(size).all()
    return {"total": total, "items": [emp_out(e) for e in items]}


@router.get("/all", response_model=list[EmployeeOut])
def list_all_employees(db: Session = Depends(get_db)):
    return [emp_out(e) for e in db.query(Employee).order_by(Employee.id).all()]


@router.get("/export/excel")
def export_employees(keyword: str = "", department_id: int | None = None,
                     status: str = "", db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    emps = build_query(db, keyword, department_id, status).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "员工信息"
    ws.append(EXPORT_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for e in emps:
        ws.append([
            e.employee_no, e.name, e.gender, e.phone, e.email,
            e.department.name if e.department else "",
            e.position.name if e.position else "",
            e.hire_date.strftime("%Y-%m-%d"), e.status,
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"employees_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/import/excel")
def import_employees(file: UploadFile = File(...), db: Session = Depends(get_db)):
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(file.file.read()))
    except Exception:
        raise HTTPException(status_code=400, detail="文件解析失败，请上传 .xlsx 格式文件")
    ws = wb.active
    success, failed = 0, []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[1]:
            continue
        no, name, gender, phone, email, dept_name, pos_name, hire_date, status = (list(row) + [""] * 9)[:9]
        try:
            if db.query(Employee).filter(Employee.employee_no == str(no)).first():
                raise ValueError("工号已存在")
            dept = db.query(Department).filter(Department.name == str(dept_name)).first()
            if not dept:
                raise ValueError(f"部门[{dept_name}]不存在")
            pos = db.query(Position).filter(Position.name == str(pos_name),
                                            Position.department_id == dept.id).first()
            if not pos:
                raise ValueError(f"职位[{pos_name}]不存在")
            if isinstance(hire_date, datetime):
                hd = hire_date.date()
            else:
                hd = datetime.strptime(str(hire_date)[:10], "%Y-%m-%d").date()
            emp = Employee(
                employee_no=str(no), name=str(name), gender=str(gender or "男"),
                phone=str(phone or ""), email=str(email or ""),
                department_id=dept.id, position_id=pos.id,
                hire_date=hd, status=str(status or "在职"),
            )
            db.add(emp)
            success += 1
        except Exception as e:
            failed.append(f"第{idx}行：{e}")
    db.commit()
    return {"message": f"导入完成：成功 {success} 条，失败 {len(failed)} 条", "failed": failed}


@router.get("/{emp_id}", response_model=EmployeeOut)
def get_employee(emp_id: int, db: Session = Depends(get_db)):
    emp = db.get(Employee, emp_id)
    if not emp:
        raise HTTPException(status_code=404, detail="员工不存在")
    return emp_out(emp)


@router.post("", response_model=EmployeeOut)
def create_employee(body: EmployeeIn, db: Session = Depends(get_db)):
    if db.query(Employee).filter(Employee.employee_no == body.employee_no).first():
        raise HTTPException(status_code=400, detail="工号已存在")
    if not db.get(Department, body.department_id):
        raise HTTPException(status_code=400, detail="部门不存在")
    if not db.get(Position, body.position_id):
        raise HTTPException(status_code=400, detail="职位不存在")
    emp = Employee(**body.model_dump())
    db.add(emp)
    db.commit()
    db.refresh(emp)
    return emp_out(emp)


@router.put("/{emp_id}", response_model=EmployeeOut)
def update_employee(emp_id: int, body: EmployeeUpdate, db: Session = Depends(get_db)):
    emp = db.get(Employee, emp_id)
    if not emp:
        raise HTTPException(status_code=404, detail="员工不存在")
    data = body.model_dump(exclude_unset=True)
    if "employee_no" in data and data["employee_no"] != emp.employee_no:
        if db.query(Employee).filter(Employee.employee_no == data["employee_no"]).first():
            raise HTTPException(status_code=400, detail="工号已存在")
    for key, value in data.items():
        setattr(emp, key, value)
    db.commit()
    db.refresh(emp)
    return emp_out(emp)


@router.delete("/{emp_id}")
def delete_employee(emp_id: int, db: Session = Depends(get_db)):
    emp = db.get(Employee, emp_id)
    if not emp:
        raise HTTPException(status_code=404, detail="员工不存在")
    if emp.user:
        raise HTTPException(status_code=400, detail="该员工已关联登录账号，请先解除关联")
    db.delete(emp)
    db.commit()
    return {"message": "删除成功"}
